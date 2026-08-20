"""Read-only workbook validation and bounded source iteration."""

from __future__ import annotations

import hashlib
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from credit_risk_er.config import SourceConfig
from credit_risk_er.models import SourceMetadata, SourceRow


class SourceContractError(ValueError):
    """Raised when the immutable source violates its declared contract."""


def sha256_file(path: Path, *, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest().upper()


def deterministic_record_id(source_sha256: str, sheet_name: str, source_row: int) -> str:
    payload = f"record-v1|{source_sha256}|{sheet_name}|{source_row}".encode()
    return hashlib.sha256(payload).hexdigest()


def validate_source(path: Path, config: SourceConfig) -> SourceMetadata:
    """Validate fingerprint and one-column schema without modifying the workbook."""
    path = path.resolve()
    if not path.is_file():
        raise SourceContractError(f"Source workbook does not exist: {path}")
    actual_hash = sha256_file(path)
    if actual_hash != config.expected_sha256.upper():
        raise SourceContractError(
            f"SHA-256 mismatch: expected {config.expected_sha256.upper()}, observed {actual_hash}"
        )

    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        sheet_names = tuple(workbook.sheetnames)
        if sheet_names != config.expected_sheet_names:
            raise SourceContractError(
                f"Workbook sheets differ: expected {config.expected_sheet_names}, observed {sheet_names}"
            )
        worksheet = workbook[config.sheet_name]
        header = worksheet.cell(row=1, column=1).value
        if header != config.column:
            raise SourceContractError(
                f"Source column mismatch: expected {config.column!r}, observed {header!r}"
            )
        if worksheet.max_column != 1:
            raise SourceContractError(
                f"Expected exactly one source column, observed {worksheet.max_column}"
            )
        max_row = worksheet.max_row
    finally:
        workbook.close()

    return SourceMetadata(
        path=path,
        sha256=actual_hash,
        size_bytes=path.stat().st_size,
        sheet_name=config.sheet_name,
        source_column=config.column,
        max_row=max_row,
    )


def iter_source_batches(metadata: SourceMetadata, batch_size: int) -> Iterator[list[SourceRow]]:
    """Yield source rows in bounded batches while preserving order and values."""
    workbook = load_workbook(metadata.path, read_only=True, data_only=False, keep_links=False)
    try:
        worksheet = workbook[metadata.sheet_name]
        batch: list[SourceRow] = []
        for source_row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2
        ):
            value = values[0]
            if value is not None and not isinstance(value, str):
                raise SourceContractError(
                    f"Unsupported cell type at {metadata.sheet_name}!A{source_row_number}: "
                    f"{type(value).__name__}"
                )
            batch.append(SourceRow(source_row_number, value))
            if len(batch) == batch_size:
                yield batch
                batch = []
        if batch:
            yield batch
    finally:
        workbook.close()
